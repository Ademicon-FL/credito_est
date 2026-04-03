"""
generator.py
Gera a planilha PrevOne com as 4 abas: BASE GRUPOS, CARTEIRA, FLUXO, RESUMO.
"""

import io
import pandas as pd
from datetime import datetime
from typing import List
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from optimizer import ResultadoOtimizacao, LinhaCarteira, gerar_fluxo_mensal


# Cores padrão
COR_HEADER = "1F3864"      # azul escuro
COR_SUBTITULO = "2F75B6"   # azul médio
COR_DESTAQUE = "D6E4F0"    # azul claro
COR_VERDE = "E2EFDA"       # verde claro (entradas)
COR_VERMELHO = "FCE4D6"    # laranja/vermelho claro (saídas)
COR_AMARELO = "FFF2CC"     # amarelo (input)
FONTE = "Arial"


def _header_style(ws, cell_ref: str, texto: str, bg=COR_HEADER, fg="FFFFFF", bold=True, size=11):
    c = ws[cell_ref]
    c.value = texto
    c.font = Font(name=FONTE, bold=bold, color=fg, size=size)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _valor(ws, cell_ref: str, valor, fmt="R$ #,##0.00", bold=False, bg=None):
    c = ws[cell_ref]
    c.value = valor
    c.number_format = fmt
    if bold:
        c.font = Font(name=FONTE, bold=True, size=10)
    else:
        c.font = Font(name=FONTE, size=10)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="right", vertical="center")


def _texto(ws, cell_ref: str, texto, bold=False, bg=None, cor="000000", size=10, wrap=False):
    c = ws[cell_ref]
    c.value = texto
    c.font = Font(name=FONTE, bold=bold, color=cor, size=size)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(vertical="center", wrap_text=wrap)


def _pct(ws, cell_ref: str, valor, decimais=2, bold=False, bg=None):
    c = ws[cell_ref]
    c.value = valor
    fmt = f'0.{"0"*decimais}%'
    c.number_format = fmt
    c.font = Font(name=FONTE, bold=bold, size=10)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="right", vertical="center")


def gerar_excel(
    resultado: ResultadoOtimizacao,
    df_base_grupos: pd.DataFrame,
    nome_cliente: str = "Cliente",
    data_base: datetime = None,
) -> bytes:
    """
    Gera o arquivo Excel PrevOne e retorna como bytes para download.
    """
    if data_base is None:
        data_base = datetime.today()

    wb = openpyxl.Workbook()

    _montar_aba_resumo(wb, resultado, nome_cliente, data_base)
    _montar_aba_carteira(wb, resultado, data_base)
    _montar_aba_fluxo(wb, resultado)
    _montar_aba_base_grupos(wb, df_base_grupos, resultado.fidc_pct)

    # Remover aba padrão
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────
# ABA RESUMO
# ─────────────────────────────────────────────────────────────
def _montar_aba_resumo(wb, resultado: ResultadoOtimizacao, nome_cliente: str, data_base: datetime):
    ws = wb.create_sheet("RESUMO")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18

    # Título
    ws.merge_cells("B1:D1")
    c = ws["B1"]
    c.value = "CRÉDITO ESTRUTURADO — CONSÓRCIO IMÓVEIS"
    c.font = Font(name=FONTE, bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COR_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("B2:D2")
    c = ws["B2"]
    c.value = nome_cliente
    c.font = Font(name=FONTE, bold=True, size=12, color=COR_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    _texto(ws, "B3", f"Data base: {data_base.strftime('%d/%m/%Y')}", size=9, cor="888888")
    fidc_label = f"{resultado.fidc_pct:.0%}" if resultado.fidc_pct > 0 else "Sem FIDC (capital próprio)"
    _texto(ws, "C3", f"FIDC: {fidc_label}", size=9, cor="888888")

    # Indicadores principais
    row = 5
    headers = ["Indicador", "Valor"]
    for col, h in enumerate(headers, 2):
        _header_style(ws, f"{get_column_letter(col)}{row}", h, bg=COR_SUBTITULO)

    dados = [
        ("Crédito líquido total", resultado.credito_liquido_total, "R$ #,##0.00", False),
        ("Crédito bruto (pré-FIDC)", resultado.credito_bruto_total, "R$ #,##0.00", False),
        ("Taxa FIDC sobre lance livre", resultado.fidc_pct, "0.00%", False) if resultado.fidc_pct > 0 else ("FIDC", "Sem FIDC", None, False),
        ("Fee FIDC total", resultado.fidc_fee_total, "R$ #,##0.00", False) if resultado.fidc_pct > 0 else None,
        ("Número de cotas", resultado.num_cotas, "0", False),
        ("Distribuição de contemplações", f"{resultado.distribuicao_meses} meses", None, False),
        ("Parcela pré-contemplação (total)", resultado.parcela_pre_total, "R$ #,##0.00", False),
        ("Parcela pós-contemplação (total)", resultado.parcela_pos_total, "R$ #,##0.00", False),
        ("TIR mensal", resultado.tir_mensal, "0.00%", True),
        ("TIR anual", resultado.tir_anual, "0.00%", True),
    ]
    dados = [d for d in dados if d]

    for i, dado in enumerate(dados):
        r = row + 1 + i
        ws.row_dimensions[r].height = 18
        label, valor, fmt, bold = dado
        _texto(ws, f"B{r}", label, bold=bold, bg="F5F5F5" if i % 2 == 0 else None, size=10)
        c = ws[f"C{r}"]
        c.value = valor
        c.font = Font(name=FONTE, bold=bold, size=10)
        c.alignment = Alignment(horizontal="right", vertical="center")
        if fmt:
            c.number_format = fmt
        if i % 2 == 0:
            c.fill = PatternFill("solid", fgColor="F5F5F5")

    # Tabela de fluxo de crédito
    row_fluxo = row + len(dados) + 3
    ws.merge_cells(f"B{row_fluxo}:D{row_fluxo}")
    _header_style(ws, f"B{row_fluxo}", "Fluxo de liberação de crédito", bg=COR_SUBTITULO)

    row_fluxo += 1
    for col, h in enumerate(["Mês", "Parcela paga (R$)", "Crédito liberado (R$)", "Crédito acumulado (R$)"], 2):
        _header_style(ws, f"{get_column_letter(col)}{row_fluxo}", h, bg=COR_HEADER, size=9)

    df_fluxo = gerar_fluxo_mensal(resultado.carteira, resultado.distribuicao_meses)
    df_fluxo_resumo = df_fluxo[df_fluxo["Crédito liberado (R$)"].notna()].head(resultado.distribuicao_meses + 2)

    for i, (_, frow) in enumerate(df_fluxo_resumo.iterrows()):
        r = row_fluxo + 1 + i
        ws.row_dimensions[r].height = 16
        bg = COR_VERDE if (frow["Crédito liberado (R$)"] or 0) > 0 else None
        _texto(ws, f"B{r}", f"Mês {int(frow['Mês'])}", bg=bg, size=10)
        _valor(ws, f"C{r}", frow["Parcela paga (R$)"], bg=bg)
        _valor(ws, f"D{r}", frow.get("Crédito liberado (R$)", 0) or 0, bg=bg)
        _valor(ws, f"E{r}", frow.get("Crédito acumulado (R$)", 0) or 0, bg=bg)

    ws.column_dimensions["E"].width = 22

    # Aviso legal
    r_aviso = row_fluxo + len(df_fluxo_resumo) + 3
    ws.merge_cells(f"B{r_aviso}:E{r_aviso}")
    c = ws[f"B{r_aviso}"]
    c.value = ("Simulação a título de ilustração. Não configura promessa ou garantia de contemplação. "
               "Percentuais de lance refletem média histórica de lances confirmados.")
    c.font = Font(name=FONTE, size=8, color="888888", italic=True)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[r_aviso].height = 28


# ─────────────────────────────────────────────────────────────
# ABA CARTEIRA
# ─────────────────────────────────────────────────────────────
def _montar_aba_carteira(wb, resultado: ResultadoOtimizacao, data_base: datetime):
    ws = wb.create_sheet("CARTEIRA")

    colunas = [
        ("Grupo", 8), ("Tranche", 8), ("Qtde Cotas", 10),
        ("Crédito (R$)", 16), ("Lance Total (R$)", 16),
        ("Lance Embutido (R$)", 18), ("Lance Livre (R$)", 16),
        ("Lance %", 9), ("Parcela Pré (R$)", 16),
        ("Nova Parcela (R$)", 16), ("Créd. Bruto/cota (R$)", 20),
        ("Taxa FIDC (R$)", 14), ("Créd. Líquido/cota (R$)", 22),
        ("Créd. Líquido Total (R$)", 24), ("TIR mensal", 10), ("Tipos de Lance", 22),
    ]
    for col, (header, width) in enumerate(colunas, 1):
        ltr = get_column_letter(col)
        ws.column_dimensions[ltr].width = width
        _header_style(ws, f"{ltr}1", header, size=9)

    formatos = [
        "0", "0", "0",
        "R$ #,##0.00", "R$ #,##0.00",
        "R$ #,##0.00", "R$ #,##0.00",
        "0.00%", "R$ #,##0.00",
        "R$ #,##0.00", "R$ #,##0.00",
        "R$ #,##0.00", "R$ #,##0.00",
        "R$ #,##0.00", "0.00%", "@",
    ]

    for i, linha in enumerate(resultado.carteira, 2):
        bg = COR_DESTAQUE if i % 2 == 0 else None
        valores = [
            linha.grupo, linha.tranche, linha.qtde_cotas,
            linha.credito, linha.lance_R,
            linha.lance_embutido_R, linha.lance_livre_R,
            linha.lance_pct, linha.parcela_lancamento,
            linha.nova_parcela, linha.ss_novo_por_cota,
            linha.fidc_fee_por_cota, linha.credito_liquido_por_cota,
            linha.credito_liquido_total, linha.custo_mensal, linha.tipos_lance,
        ]
        for col, (val, fmt) in enumerate(zip(valores, formatos), 1):
            c = ws[f"{get_column_letter(col)}{i}"]
            c.value = val
            c.number_format = fmt
            c.font = Font(name=FONTE, size=9)
            c.alignment = Alignment(horizontal="right" if fmt != "@" else "left", vertical="center")
            if bg:
                c.fill = PatternFill("solid", fgColor=bg)

    # Linha de totais
    r_total = len(resultado.carteira) + 2
    _header_style(ws, f"A{r_total}", "TOTAL", bg=COR_HEADER, size=9)
    _valor(ws, f"C{r_total}", resultado.num_cotas, "0", bold=True, bg=COR_HEADER)

    for col, val in [
        (14, resultado.credito_liquido_total),
        (12, resultado.fidc_fee_total),
    ]:
        c = ws[f"{get_column_letter(col)}{r_total}"]
        c.value = val
        c.number_format = "R$ #,##0.00"
        c.font = Font(name=FONTE, bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor=COR_HEADER)
        c.alignment = Alignment(horizontal="right", vertical="center")


# ─────────────────────────────────────────────────────────────
# ABA FLUXO
# ─────────────────────────────────────────────────────────────
def _montar_aba_fluxo(wb, resultado: ResultadoOtimizacao):
    ws = wb.create_sheet("FLUXO")

    colunas = [
        ("Mês", 6), ("Cotas contempl.", 14), ("Parcela paga (R$)", 18),
        ("Crédito liberado (R$)", 20), ("Crédito acumulado (R$)", 22), ("Caixa (R$)", 16),
    ]
    for col, (header, width) in enumerate(colunas, 1):
        ltr = get_column_letter(col)
        ws.column_dimensions[ltr].width = width
        _header_style(ws, f"{ltr}1", header, size=9)

    formatos = ["0", "0", "R$ #,##0.00", "R$ #,##0.00", "R$ #,##0.00", "R$ #,##0.00"]

    df_fluxo = gerar_fluxo_mensal(resultado.carteira, resultado.distribuicao_meses)

    for i, (_, row) in enumerate(df_fluxo.iterrows(), 2):
        credito_mes = row.get("Crédito liberado (R$)", 0) or 0
        bg = COR_VERDE if credito_mes > 0 else (COR_VERMELHO if (row.get("Caixa (R$)", 0) or 0) < -1000 else None)

        valores = [
            row["Mês"],
            row.get("Cotas contempladas", None),
            row.get("Parcela paga (R$)", 0),
            credito_mes if credito_mes > 0 else None,
            row.get("Crédito acumulado (R$)", None),
            row.get("Caixa (R$)", 0),
        ]
        for col, (val, fmt) in enumerate(zip(valores, formatos), 1):
            c = ws[f"{get_column_letter(col)}{i}"]
            c.value = val
            c.number_format = fmt
            c.font = Font(name=FONTE, size=9)
            c.alignment = Alignment(horizontal="right", vertical="center")
            if bg:
                c.fill = PatternFill("solid", fgColor=bg)

    # TIR no topo
    ws["H1"] = "TIR mensal"
    ws["H2"] = resultado.tir_mensal
    ws["H2"].number_format = "0.0000%"
    ws["I1"] = "TIR anual"
    ws["I2"] = resultado.tir_anual
    ws["I2"].number_format = "0.00%"
    for ref in ["H1", "I1"]:
        ws[ref].font = Font(name=FONTE, bold=True, size=9)
    for ref in ["H2", "I2"]:
        ws[ref].font = Font(name=FONTE, bold=True, size=11, color=COR_HEADER)
        ws[ref].alignment = Alignment(horizontal="right")


# ─────────────────────────────────────────────────────────────
# ABA BASE GRUPOS
# ─────────────────────────────────────────────────────────────
def _montar_aba_base_grupos(wb, df_base: pd.DataFrame, fidc_pct: float):
    ws = wb.create_sheet("BASE GRUPOS")

    colunas_map = {
        "grupo": ("Grupo", 8, "0"),
        "contemp_ll_mes": ("Contempl./mês LL", 16, "0.0"),
        "participantes": ("Part.", 8, "0"),
        "credito": ("Crédito (R$)", 16, "R$ #,##0.00"),
        "prazo_original": ("Prazo Orig.", 10, "0"),
        "taxa_adm": ("Taxa Adm.", 10, "0.00%"),
        "prazo_atual": ("Prazo Atual", 10, "0"),
        "parcela_lancamento": ("Parcela (R$)", 14, "R$ #,##0.00"),
        "lance_total_parcelas": ("Lance (parc.)", 12, "0.0"),
        "lance_R": ("Lance R$", 14, "R$ #,##0.00"),
        "lance_pct": ("Lance %", 9, "0.00%"),
        "ss_novo": ("Créd. Bruto (R$)", 16, "R$ #,##0.00"),
        "fidc_fee": (f"Fee FIDC {fidc_pct:.0%} (R$)", 16, "R$ #,##0.00"),
        "credito_liquido": ("Créd. Líquido (R$)", 18, "R$ #,##0.00"),
        "nova_parcela": ("Nova Parcela (R$)", 16, "R$ #,##0.00"),
        "custo_mensal": ("Custo Mensal", 12, "0.0000%"),
        "custo_anual": ("Custo Anual", 12, "0.00%"),
        "tipos_lance": ("Tipos de Lance", 24, "@"),
    }

    colunas = [c for c in colunas_map if c in df_base.columns]

    for col, campo in enumerate(colunas, 1):
        ltr = get_column_letter(col)
        header, width, fmt = colunas_map[campo]
        ws.column_dimensions[ltr].width = width
        _header_style(ws, f"{ltr}1", header, size=9)

    for i, (_, row) in enumerate(df_base.iterrows(), 2):
        bg = "F5F5F5" if i % 2 == 0 else None
        for col, campo in enumerate(colunas, 1):
            val = row.get(campo, None)
            if pd.isna(val) if not isinstance(val, str) else False:
                val = None
            c = ws[f"{get_column_letter(col)}{i}"]
            c.value = val
            c.number_format = colunas_map[campo][2]
            c.font = Font(name=FONTE, size=9)
            c.alignment = Alignment(
                horizontal="right" if colunas_map[campo][2] != "@" else "left",
                vertical="center",
            )
            if bg:
                c.fill = PatternFill("solid", fgColor=bg)
